"""Geração do XLSX redline com marcações por célula e aba Summary."""
from __future__ import annotations

import io
import logging
import zipfile
from typing import Dict, List, Optional, Set, Tuple
from xml.etree import ElementTree as ET

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from app.xlsx.compare import CellDiff, SheetDiff, XlsxDiff, compare_xlsx
from app.xlsx.models import ColorConfig

logger = logging.getLogger(__name__)

_RELS_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_PKG_NS = "http://schemas.openxmlformats.org/package/2006/content-types"

# Partes do XLSX base que o openpyxl costuma dropar no save e que, se
# ficarem pela metade, fazem o Excel abrir em modo recuperação.
_PRESERVE_DIR_PREFIXES = (
    "xl/media/",
    "xl/printerSettings/",
    "xl/webextensions/",
    "xl/threadedComments/",
    "xl/persons/",
    "customXml/",
)
_PRESERVE_EXACT = (
    "xl/connections.xml",
    "xl/comments1.xml",
    "xl/drawings/vmlDrawing1.vml",
    "xl/calcChain.xml",
)

def _hex_rgb(color: str) -> str:
    return color.lstrip("#").upper()


def _rich_text_for_modified(
    base_text: str, compare_text: str, colors: ColorConfig
) -> CellRichText:
    deletion_font = InlineFont(color=_hex_rgb(colors.deletion), strike=True)
    insertion_font = InlineFont(color=_hex_rgb(colors.insertion), u="single")
    return CellRichText(
        TextBlock(deletion_font, base_text),
        TextBlock(insertion_font, compare_text),
    )


def generate_redline_xlsx(
    base_bytes: bytes,
    compare_bytes: bytes,
    colors: Optional[ColorConfig] = None,
    base_filename: str = "",
    compare_filename: str = "",
) -> Tuple[bytes, XlsxDiff]:
    """Gera bytes do XLSX redline e o ``XlsxDiff`` usado (para stats do job).

    Parte do arquivo **revisado** (como o Word Compare): não risca abas
    inteiras removidas, não insere/apaga linhas (isso quebrava fórmulas e
    abria o Excel em recuperação). Mudanças de célula são só estilo/valor
    pontual; abas só no base aparecem no Compare Summary.
    """
    colors = colors or ColorConfig()
    diff = compare_xlsx(base_bytes, compare_bytes, base_filename, compare_filename)

    # Canvas = revisado (documento atual).
    out_wb: Workbook = load_workbook(io.BytesIO(compare_bytes), data_only=False)
    base_wb: Workbook = load_workbook(io.BytesIO(base_bytes), data_only=False)
    compare_wb: Workbook = load_workbook(io.BytesIO(compare_bytes), data_only=False)

    try:
        for sheet_diff in diff.sheets:
            if sheet_diff.base_only:
                # Aba sumiu no revisado — não copiar riscata (corrompe o pacote
                # e pinta a tela de vermelho). Só reportar no summary.
                continue

            if sheet_diff.compare_only:
                if sheet_diff.name in out_wb.sheetnames:
                    _add_sheet_banner(
                        out_wb[sheet_diff.name],
                        "Aba nova (não existia no arquivo base).",
                        colors,
                        insertion=True,
                    )
                continue

            if sheet_diff.name not in out_wb.sheetnames:
                continue
            base_ws = (
                base_wb[sheet_diff.name]
                if sheet_diff.name in base_wb.sheetnames
                else None
            )
            _mark_compare_sheet_from_diff(
                out_wb[sheet_diff.name],
                base_ws,
                compare_wb[sheet_diff.name],
                sheet_diff,
                colors,
            )

        _append_summary_sheet(out_wb, diff, colors)

        buffer = io.BytesIO()
        out_wb.save(buffer)
        raw = buffer.getvalue()
        try:
            # Preservar mídia/estrutura a partir do revisado (canvas).
            raw = _finalize_xlsx_package(compare_bytes, raw)
        except Exception:
            logger.exception(
                "Falha ao preservar mídia/partes do XLSX revisado; "
                "seguindo com o arquivo gerado pelo openpyxl."
            )
        return raw, diff
    finally:
        out_wb.close()
        base_wb.close()
        compare_wb.close()


def _finalize_xlsx_package(base_bytes: bytes, redline_bytes: bytes) -> bytes:
    """Reinjeta mídia e remove links externos quebrados após o save do openpyxl.

    O round-trip do openpyxl descarta imagens grandes, query tables e
    connections, mas deixa relacionamentos/externalLinks pela metade — o
    Excel então abre em recuperação e “remove” intervalos de dados externos.
    """
    base_zip = zipfile.ZipFile(io.BytesIO(base_bytes))
    red_zip = zipfile.ZipFile(io.BytesIO(redline_bytes))
    out_buf = io.BytesIO()
    base_names = set(base_zip.namelist())

    preserve: Dict[str, bytes] = {}
    for name in base_names:
        if name in _PRESERVE_EXACT or any(name.startswith(p) for p in _PRESERVE_DIR_PREFIXES):
            preserve[name] = base_zip.read(name)
        if name.startswith("xl/queryTables/"):
            preserve[name] = base_zip.read(name)

    with zipfile.ZipFile(out_buf, "w", compression=zipfile.ZIP_DEFLATED) as out_zip:
        written: Set[str] = set()
        for info in red_zip.infolist():
            name = info.filename
            # openpyxl deixa externalLinks inconsistentes → Excel em recuperação.
            if name.startswith("xl/externalLinks/"):
                continue
            # Mídia: usar sempre a do base (completa).
            if name.startswith("xl/media/"):
                continue
            data = red_zip.read(name)
            if name == "xl/_rels/workbook.xml.rels":
                data = _strip_external_link_rels(data)
            elif name == "xl/workbook.xml":
                data = _strip_external_references(data)
            elif name == "[Content_Types].xml":
                data = _merge_content_types(data, set(preserve.keys()))
            elif name.startswith("xl/worksheets/_rels/"):
                data = _merge_sheet_rels(data, name, base_zip, base_names)
            out_zip.writestr(name, data)
            written.add(name)

        for name, data in preserve.items():
            if name.startswith("xl/externalLinks/"):
                continue
            if name not in written:
                out_zip.writestr(name, data)
                written.add(name)

    base_zip.close()
    red_zip.close()
    logger.info(
        "XLSX finalizado: reinjetadas %d partes do base; links externos removidos.",
        sum(1 for n in written if n in preserve),
    )
    return out_buf.getvalue()


def _strip_external_link_rels(rels_xml: bytes) -> bytes:
    try:
        root = ET.fromstring(rels_xml)
    except ET.ParseError:
        return rels_xml
    ns = {"r": _RELS_NS}
    removed = False
    for rel in list(root.findall("r:Relationship", ns)):
        rel_type = rel.attrib.get("Type", "")
        target = rel.attrib.get("Target", "")
        if "externalLink" in rel_type or "externalLinks/" in target:
            root.remove(rel)
            removed = True
    if not removed:
        return rels_xml
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _strip_external_references(workbook_xml: bytes) -> bytes:
    try:
        root = ET.fromstring(workbook_xml)
    except ET.ParseError:
        return workbook_xml
    for child in list(root):
        tag = child.tag.rsplit("}", 1)[-1]
        if tag == "externalReferences":
            root.remove(child)
            return ET.tostring(root, encoding="utf-8", xml_declaration=True)
    return workbook_xml


def _merge_content_types(ct_xml: bytes, preserved_names: Set[str]) -> bytes:
    try:
        root = ET.fromstring(ct_xml)
    except ET.ParseError:
        return ct_xml
    existing = {
        node.attrib.get("PartName", "")
        for node in root
        if node.tag.endswith("Override")
    }
    # Remove overrides de externalLinks (partes não serão gravadas).
    for node in list(root):
        if node.tag.endswith("Override"):
            part = node.attrib.get("PartName", "")
            if "externalLink" in part:
                root.remove(node)
                existing.discard(part)

    type_for = {
        "xl/connections.xml": "application/vnd.openxmlformats-officedocument.spreadsheetml.connections+xml",
        "xl/comments1.xml": "application/vnd.openxmlformats-officedocument.spreadsheetml.comments+xml",
        "xl/calcChain.xml": "application/vnd.openxmlformats-officedocument.spreadsheetml.calcChain+xml",
    }
    for name in preserved_names:
        if name.startswith("xl/externalLinks/"):
            continue
        part = "/" + name
        if name in type_for and part not in existing:
            ET.SubElement(
                root, "{%s}Override" % _PKG_NS, PartName=part, ContentType=type_for[name]
            )
            existing.add(part)
        elif name.startswith("xl/queryTables/") and name.endswith(".xml") and part not in existing:
            ET.SubElement(
                root,
                "{%s}Override" % _PKG_NS,
                PartName=part,
                ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.queryTable+xml",
            )
            existing.add(part)
        elif name.startswith("xl/printerSettings/") and part not in existing:
            ET.SubElement(
                root,
                "{%s}Override" % _PKG_NS,
                PartName=part,
                ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.printerSettings",
            )
            existing.add(part)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _merge_sheet_rels(
    rels_xml: bytes, rels_path: str, base_zip: zipfile.ZipFile, base_names: Set[str]
) -> bytes:
    """Reanexa relacionamentos de imagem/vml/comments do base que o openpyxl omitiu."""
    if rels_path not in base_names:
        return rels_xml
    try:
        red_root = ET.fromstring(rels_xml)
        base_root = ET.fromstring(base_zip.read(rels_path))
    except ET.ParseError:
        return rels_xml

    def _targets(root: ET.Element) -> Dict[str, ET.Element]:
        out: Dict[str, ET.Element] = {}
        for rel in root:
            if not rel.tag.endswith("Relationship"):
                continue
            target = rel.attrib.get("Target", "")
            out[target] = rel
        return out

    red_by_target = _targets(red_root)
    used_ids = {
        rel.attrib.get("Id", "")
        for rel in red_root
        if rel.tag.endswith("Relationship")
    }
    next_id = 1

    def _alloc_id() -> str:
        nonlocal next_id
        while "rId%d" % next_id in used_ids:
            next_id += 1
        rid = "rId%d" % next_id
        used_ids.add(rid)
        next_id += 1
        return rid

    keep_types = ("image", "vmlDrawing", "comments", "threadedComment")
    for target, rel in _targets(base_root).items():
        rel_type = rel.attrib.get("Type", "")
        if not any(k in rel_type for k in keep_types):
            continue
        if target in red_by_target:
            continue
        new_rel = ET.SubElement(red_root, "{%s}Relationship" % _RELS_NS)
        new_rel.set("Id", _alloc_id())
        new_rel.set("Type", rel_type)
        new_rel.set("Target", target)

    return ET.tostring(red_root, encoding="utf-8", xml_declaration=True)
def _deletion_font(colors: ColorConfig) -> Font:
    return Font(color=_hex_rgb(colors.deletion), strike=True)


def _insertion_font(colors: ColorConfig) -> Font:
    return Font(color=_hex_rgb(colors.insertion), underline="single")


def _apply_font_preserving_style(cell: Cell, new_font: Font) -> None:
    try:
        existing = cell.font
        cell.font = Font(
            name=existing.name,
            size=existing.size,
            bold=existing.bold,
            italic=new_font.italic if new_font.italic is not None else existing.italic,
            color=new_font.color or existing.color,
            strike=new_font.strike if new_font.strike is not None else existing.strike,
            underline=new_font.underline or existing.underline,
            family=existing.family,
            scheme=existing.scheme,
        )
    except Exception:
        cell.font = new_font


def _add_sheet_banner(
    ws: Worksheet, text: str, colors: ColorConfig, insertion: bool = True
) -> None:
    """Insere aviso na linha 1 sem deslocar o restante (escreve em ZZ1)."""
    # Evita insert_rows — usa célula longe à direita no topo.
    cell = ws.cell(row=1, column=50, value=text)
    cell.font = _insertion_font(colors) if insertion else _deletion_font(colors)
    cell.fill = PatternFill(
        start_color="DBEAFE" if insertion else "FEE2E2",
        end_color="DBEAFE" if insertion else "FEE2E2",
        fill_type="solid",
    )


def _mark_compare_sheet_from_diff(
    out_ws: Worksheet,
    base_ws: Optional[Worksheet],
    compare_ws: Worksheet,
    sheet_diff: SheetDiff,
    colors: ColorConfig,
) -> None:
    """Marca mudanças na aba do revisado, sem insert/delete de linhas."""
    del_font = _deletion_font(colors)
    ins_font = _insertion_font(colors)

    for row_diff in sheet_diff.rows:
        if row_diff.status == "delete":
            # Linha só existia no base — não está no revisado; só no summary.
            continue

        if row_diff.compare_row_index is None:
            continue

        sheet_row = row_diff.compare_row_index + 1

        if row_diff.status == "insert":
            for cell_diff in row_diff.cells:
                if not cell_diff.compare_cell or not cell_diff.compare_cell.value:
                    continue
                col = cell_diff.compare_cell.col
                out_cell = out_ws.cell(row=sheet_row, column=col)
                if out_cell.value not in (None, ""):
                    _apply_font_preserving_style(out_cell, ins_font)
            continue

        if row_diff.status == "equal":
            continue

        # replace: só células diferentes
        for cell_diff in row_diff.cells:
            if cell_diff.status == "equal":
                continue
            col = (
                cell_diff.compare_cell.col
                if cell_diff.compare_cell is not None
                else cell_diff.col
            )
            out_cell = out_ws.cell(row=sheet_row, column=col)
            _write_cell_mark_on_compare(
                out_cell,
                cell_diff,
                base_ws,
                compare_ws,
                del_font,
                ins_font,
                colors,
            )


def _write_cell_mark_on_compare(
    out_cell: Cell,
    diff: CellDiff,
    base_ws: Optional[Worksheet],
    compare_ws: Worksheet,
    del_font: Font,
    ins_font: Font,
    colors: ColorConfig,
) -> None:
    """Aplica marca visual na célula do revisado sem destruir fórmulas."""
    base_cell_src = None
    compare_cell_src = None
    if diff.base_cell is not None and base_ws is not None:
        base_cell_src = base_ws.cell(row=diff.base_cell.row, column=diff.base_cell.col)
    if diff.compare_cell is not None:
        compare_cell_src = compare_ws.cell(
            row=diff.compare_cell.row, column=diff.compare_cell.col
        )

    base_text = diff.base_cell.value if diff.base_cell else ""
    compare_text = diff.compare_cell.value if diff.compare_cell else ""
    compare_is_formula = bool(
        compare_cell_src is not None
        and isinstance(compare_cell_src.value, str)
        and str(compare_cell_src.value).startswith("=")
    )

    if diff.status == "insert":
        if out_cell.value not in (None, ""):
            _apply_font_preserving_style(out_cell, ins_font)
        return

    if diff.status == "delete":
        # Célula esvaziada no revisado: se ainda há valor residual, risca;
        # senão só fundo.
        if out_cell.value not in (None, ""):
            _apply_font_preserving_style(out_cell, del_font)
        out_cell.fill = _empty_cell_fill(colors)
        return

    # modified
    if compare_is_formula:
        # Mantém a fórmula; só destaca.
        _apply_font_preserving_style(out_cell, ins_font)
        out_cell.fill = PatternFill(
            start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"
        )
        return

    if base_text and compare_text and base_text != compare_text:
        # Mostra antigo riscato + novo sublinhado (texto, não fórmula).
        out_cell.value = _rich_text_for_modified(base_text, compare_text, colors)
        return

    if compare_text and not base_text:
        _apply_font_preserving_style(out_cell, ins_font)
    elif base_text and not compare_text:
        out_cell.fill = _empty_cell_fill(colors)
    else:
        _apply_font_preserving_style(out_cell, ins_font)


def _empty_cell_fill(colors: ColorConfig) -> PatternFill:
    hex_color = _hex_rgb(colors.deletion)
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    pale_hex = "%02X%02X%02X" % (
        int(r * 0.12 + 255 * 0.88),
        int(g * 0.12 + 255 * 0.88),
        int(b * 0.12 + 255 * 0.88),
    )
    return PatternFill(start_color=pale_hex, end_color=pale_hex, fill_type="solid")


def _append_summary_sheet(wb: Workbook, diff: XlsxDiff, colors: ColorConfig) -> None:
    # Sempre "Compare Summary" — não reutilizar o nome "Summary" (comum em Cap Tables).
    name = "Compare Summary"
    suffix = 1
    while name in wb.sheetnames:
        suffix += 1
        name = "Compare Summary %d" % suffix

    ws = wb.create_sheet(title=name)
    stats = diff.stats

    ws["A1"] = "Smart Compare Report"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:B1")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = "Compared on: %s" % diff.summary.compared_on
    ws["A2"].font = Font(italic=True, color="707070")
    ws.merge_cells("A2:B2")
    ws["A2"].alignment = Alignment(horizontal="center")

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    row = 4
    for label, value in [
        ("Base file:", diff.base_filename or "—"),
        ("Modified file:", diff.compare_filename or "—"),
        ("Value Change:", str(stats.value_changes)),
        ("Modified:", str(stats.modified_cells)),
    ]:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        for col in (1, 2):
            ws.cell(row=row, column=col).border = thin_border
        row += 1

    ins_color = _hex_rgb(colors.insertion)
    del_color = _hex_rgb(colors.deletion)
    ins_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    del_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    yellow_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")

    ins_label_font = Font(color=ins_color, bold=True, underline="single")
    del_label_font = Font(color=del_color, bold=True, strike=True)
    ins_value_font = Font(color=ins_color, bold=True, underline="single")
    del_value_font = Font(color=del_color, bold=True, strike=True)
    plain_bold = Font(bold=True)

    detail_rows = [
        ("Row Add:", str(stats.row_add), ins_fill, ins_label_font, ins_value_font),
        ("Row Del:", str(stats.row_del), del_fill, del_label_font, del_value_font),
        ("Col Add:", str(stats.col_add), ins_fill, ins_label_font, ins_value_font),
        ("Col Del:", str(stats.col_del), del_fill, del_label_font, del_value_font),
        (
            "Formula And Value Change:",
            str(stats.formula_and_value_changes),
            yellow_fill,
            Font(bold=True, color="713F12"),
            plain_bold,
        ),
        (
            "Formula Change Only:",
            str(stats.formula_only_changes),
            green_fill,
            Font(bold=True, color="166534", underline="single"),
            plain_bold,
        ),
        ("Emptied Cells:", str(stats.emptied_cells), del_fill, del_label_font, del_value_font),
    ]

    for label, value, fill, label_font, value_font in detail_rows:
        label_cell = ws.cell(row=row, column=1, value=label)
        value_cell = ws.cell(row=row, column=2, value=value)
        label_cell.font = label_font
        value_cell.font = value_font
        for cell in (label_cell, value_cell):
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = (
                Alignment(horizontal="left", indent=1)
                if cell is label_cell
                else Alignment(horizontal="center")
            )
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Change Type").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Number of changes").font = Font(bold=True)
    header_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    for col in (1, 2):
        ws.cell(row=row, column=col).fill = header_fill
        ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
    row += 1

    for label, value, color, strike, underline in [
        ("Deletions", str(diff.summary.deletions), colors.deletion, True, False),
        ("Insertions", str(diff.summary.insertions), colors.insertion, False, True),
        ("Moved", str(diff.summary.moved), colors.moved, False, False),
    ]:
        label_cell = ws.cell(row=row, column=1, value=label)
        value_cell = ws.cell(row=row, column=2, value=value)
        label_cell.font = Font(
            color=_hex_rgb(color),
            bold=True,
            strike=strike,
            underline="single" if underline else None,
        )
        value_cell.font = Font(color=_hex_rgb(color), bold=True)
        for cell in (label_cell, value_cell):
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Sheets changed:").font = Font(bold=True)
    row += 1

    for sheet_diff in diff.sheets:
        if sheet_diff.name == name:
            continue
        if sheet_diff.compare_only:
            status = "aba nova no revisado"
        elif sheet_diff.base_only:
            status = "aba removida (só no base — ver detalhes abaixo)"
        else:
            status = "%d mudanças de célula/linha" % _count_sheet_changes(sheet_diff)
        ws.cell(row=row, column=1, value=sheet_diff.name)
        ws.cell(row=row, column=2, value=status)
        row += 1

    # Lista abas removidas com destaque
    removed = [s for s in diff.sheets if s.base_only]
    if removed:
        row += 1
        ws.cell(row=row, column=1, value="Abas presentes só no base (não copiadas para o redline):").font = Font(bold=True, color=_hex_rgb(colors.deletion))
        row += 1
        for sheet_diff in removed:
            ws.cell(row=row, column=1, value=sheet_diff.name).font = Font(
                color=_hex_rgb(colors.deletion), strike=True
            )
            ws.cell(row=row, column=2, value="%d linhas no base" % len(sheet_diff.rows))
            row += 1

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 34

def _count_sheet_changes(sheet_diff: SheetDiff) -> int:
    count = 0
    for row in sheet_diff.rows:
        if row.status in ("insert", "delete"):
            count += 1
        elif row.status == "replace":
            count += sum(1 for c in row.cells if c.status != "equal")
    return count
