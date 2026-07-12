"""Leitura e normalização de células XLSX para comparação."""
from __future__ import annotations

import io
import logging
from dataclasses import dataclass, field
from typing import List, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Espelho de app.extract.xlsx_extractor — evita varrer colunas fantasma (ex. 16383).
_MAX_XLSX_COLS = 256


@dataclass
class XlsxCell:
    row: int
    col: int
    coordinate: str
    value: str
    is_formula: bool = False


@dataclass
class XlsxSheet:
    name: str
    rows: List[List[XlsxCell]] = field(default_factory=list)
    max_row: int = 0
    max_col: int = 0

    def row_text(self, row_idx: int) -> str:
        if 0 <= row_idx < len(self.rows):
            return "\t".join(cell.value for cell in self.rows[row_idx])
        return ""

    def cell_at(self, row_idx: int, col_idx: int) -> Optional[XlsxCell]:
        if 0 <= row_idx < len(self.rows) and 0 <= col_idx < len(self.rows[row_idx]):
            return self.rows[row_idx][col_idx]
        return None


def _cell_value_to_str(value: object, is_formula: bool) -> str:
    if value is None:
        return ""
    if is_formula:
        return str(value)
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return "%.10g" % value
    return str(value)


def _trim_sheet_columns(sheet: XlsxSheet) -> None:
    """Corta colunas vazias à direita e atualiza max_col."""
    if not sheet.rows:
        sheet.max_col = 0
        return
    last = -1
    for row in sheet.rows:
        for i in range(len(row) - 1, -1, -1):
            if row[i].value:
                last = max(last, i)
                break
    keep = last + 1
    if keep <= 0:
        sheet.rows = []
        sheet.max_row = 0
        sheet.max_col = 0
        return
    if keep < max(len(r) for r in sheet.rows):
        sheet.rows = [r[:keep] for r in sheet.rows]
    sheet.max_col = keep
    sheet.max_row = len(sheet.rows)


def extract_sheets(file_bytes: bytes) -> List[XlsxSheet]:
    """Parse um XLSX em estruturas por aba (fórmulas preservadas)."""
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=False, read_only=False)
    sheets: List[XlsxSheet] = []

    try:
        for ws in workbook.worksheets:
            declared_cols = int(ws.max_column or 0)
            max_col = min(declared_cols, _MAX_XLSX_COLS) if declared_cols else 0
            if declared_cols > _MAX_XLSX_COLS:
                logger.warning(
                    "Aba '%s' declara %d colunas; limitando a %d (dimensão fantasma?).",
                    ws.title, declared_cols, _MAX_XLSX_COLS,
                )
            max_row = int(ws.max_row or 0)
            sheet = XlsxSheet(
                name=ws.title,
                max_row=max_row,
                max_col=max_col,
            )
            for row_idx in range(1, max_row + 1):
                row_cells: List[XlsxCell] = []
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = cell.value
                    is_formula = isinstance(value, str) and value.startswith("=")
                    text_value = _cell_value_to_str(value, is_formula)
                    row_cells.append(
                        XlsxCell(
                            row=row_idx,
                            col=col_idx,
                            coordinate="%s%d" % (get_column_letter(col_idx), row_idx),
                            value=text_value,
                            is_formula=is_formula,
                        )
                    )
                sheet.rows.append(row_cells)
            _trim_sheet_columns(sheet)
            sheets.append(sheet)
    finally:
        workbook.close()

    return sheets


def sheet_row_fingerprints(sheet: XlsxSheet) -> List[str]:
    return [sheet.row_text(i) for i in range(len(sheet.rows))]
