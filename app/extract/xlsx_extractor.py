"""Extração de planilhas Excel (.xlsx) para o modelo canônico.

Cada aba vira um título seguido de um bloco de tabela com todas as linhas
preenchidas. Valores são lidos com ``data_only=True`` (fórmulas → resultado).

Planilhas com ``max_column`` fantasma (ex.: 16383 por célula/estilo residual)
são cortadas ao conteúdo real — senão o compare trava em milhões de células.
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, List, Optional, Sequence

from openpyxl import load_workbook

from app.models import Block, BlockKind, Cell, Document, Run

logger = logging.getLogger(__name__)

# Cap tables / orçamentos raramente passam disso; acima disso é quase sempre
# dimensão fantasma do Excel (estilo até a coluna XFD).
_MAX_XLSX_COLS = 256


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def _row_has_content(cells: List[Cell]) -> bool:
    return any(c.text for c in cells)


def _trim_trailing_empty_columns(rows: List[List[Cell]]) -> List[List[Cell]]:
    if not rows:
        return rows
    width = max(len(r) for r in rows)
    last = -1
    for r in rows:
        for i in range(len(r) - 1, -1, -1):
            if r[i].text:
                last = max(last, i)
                break
    if last < 0:
        return []
    keep = last + 1
    if keep >= width:
        return rows
    return [r[:keep] for r in rows]


def extract_xlsx(path: str) -> Document:
    """Carrega um .xlsx e normaliza cada aba como título + tabela."""
    p = Path(path)
    wb = load_workbook(filename=str(p), read_only=True, data_only=True)
    blocks: List[Block] = []
    sheet_count = len(wb.sheetnames)

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            blocks.append(
                Block(
                    kind=BlockKind.HEADING,
                    runs=[Run(text=sheet_name)],
                    level=1,
                )
            )
            declared = int(ws.max_column or 0)
            max_col: Optional[int] = min(declared, _MAX_XLSX_COLS) if declared else _MAX_XLSX_COLS
            if declared > _MAX_XLSX_COLS:
                logger.warning(
                    "Aba '%s' declara %d colunas; limitando a %d (dimensão fantasma?).",
                    sheet_name, declared, _MAX_XLSX_COLS,
                )
            rows: List[List[Cell]] = []
            for row in ws.iter_rows(max_col=max_col, values_only=True):
                values: Sequence[Any] = row or ()
                cells = [Cell(runs=[Run(text=_cell_text(v))]) for v in values]
                if _row_has_content(cells):
                    rows.append(cells)
            rows = _trim_trailing_empty_columns(rows)
            if rows:
                blocks.append(Block(kind=BlockKind.TABLE, rows=rows))
    finally:
        wb.close()

    logger.info("Extraídas %d aba(s) de %s", sheet_count, p.name)
    return Document(source_path=str(p), fmt="xlsx", title=p.stem, blocks=blocks)
