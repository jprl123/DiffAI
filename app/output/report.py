"""Relatórios analíticos do Compare-Docs: JSON, XLSX e HTML.

- write_json_report: dump direto de result.to_dict().
- write_xlsx_report: planilha com aba "Mudanças" e aba "Estatísticas".
- write_html_report: arquivo único auto-contido (CSS+JS inline, zero CDN),
  pt-BR, com abas, filtros, busca, diff lado a lado e tema claro/escuro.
"""
from __future__ import annotations

import json
import logging
import os
import re
from enum import Enum
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models import Change, ComparisonResult

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Traduções pt-BR
# ---------------------------------------------------------------------------

TYPE_LABELS_PT: Dict[str, str] = {
    "equal": "Igual",
    "insert": "Inserção",
    "delete": "Exclusão",
    "modify": "Modificação",
    "move": "Movimentação",
    "move_modify": "Movimentação com modificação",
}

CATEGORY_LABELS_PT: Dict[str, str] = {
    "content": "Conteúdo",
    "formatting": "Formatação",
    "noise_date": "Rotineira (data)",
    "noise_version": "Rotineira (versão)",
    "noise_pagenum": "Rotineira (página)",
    "noise_whitespace": "Rotineira (espaçamento)",
    "noise_punct": "Rotineira (pontuação)",
    "table": "Tabela",
    "image": "Imagem",
    "metadata": "Metadados",
}

_ILLEGAL_XLSX_CHARS = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
_XLSX_MAX_CELL = 32767


def _enum_value(value: Any) -> str:
    if isinstance(value, Enum):
        return str(value.value)
    return str(value) if value is not None else ""


def _type_label(change_type: Any) -> str:
    key = _enum_value(change_type)
    return TYPE_LABELS_PT.get(key, key or "—")


def _category_label(category: Any) -> str:
    key = _enum_value(category)
    return CATEGORY_LABELS_PT.get(key, key or "—")


def _validate(result: ComparisonResult, out_path: str, what: str) -> None:
    if result is None:
        raise ValueError(
            "Resultado de comparação inválido: recebido None ao gerar %s." % what
        )
    if getattr(result, "changes", None) is None or getattr(result, "stats", None) is None:
        raise ValueError(
            "Resultado de comparação inválido: estrutura ComparisonResult "
            "incompleta (changes/stats ausentes) ao gerar %s." % what
        )
    if not out_path or not str(out_path).strip():
        raise ValueError("Caminho de saída do relatório %s não informado." % what)


def _ensure_dir(out_path: str) -> None:
    out_dir = os.path.dirname(os.path.abspath(str(out_path)))
    try:
        os.makedirs(out_dir, exist_ok=True)
    except OSError as exc:
        raise ValueError(
            "Não foi possível criar o diretório de saída '%s': %s" % (out_dir, exc)
        )


def _page_of(change: Change) -> Optional[int]:
    if getattr(change, "page_compare", None) is not None:
        return change.page_compare
    return getattr(change, "page_base", None)


def _breadcrumb(section_path: Optional[List[str]]) -> str:
    if not section_path:
        return ""
    return " › ".join(str(s) for s in section_path if s)


def _stats_pairs(result: ComparisonResult) -> List[tuple]:
    stats = result.stats
    pairs = [
        ("Data da comparação", result.compared_at or "—"),
        ("Arquivo base", os.path.basename(result.base_path) or "—"),
        ("Arquivo revisado", os.path.basename(result.compare_path) or "—"),
        ("Total de alterações", stats.total_changes),
        ("Inserções", stats.insertions),
        ("Exclusões", stats.deletions),
        ("Modificações", stats.modifications),
        ("Movimentações", stats.moves),
        ("Mudanças de conteúdo", stats.content_changes),
        ("Mudanças de formatação", stats.formatting_changes),
        ("Mudanças rotineiras (ruído)", stats.noise_changes),
        ("Mudanças em tabelas", stats.table_changes),
        ("Mudanças em imagens", stats.image_changes),
        (
            "Páginas alteradas",
            ", ".join(str(p) for p in (stats.changed_pages or [])) or "—",
        ),
        ("Duração (s)", round(getattr(result, "duration_seconds", 0.0) or 0.0, 2)),
    ]
    for key, count in sorted((stats.by_category or {}).items()):
        pairs.append(("Por categoria — %s" % _category_label(key), count))
    return pairs


# ---------------------------------------------------------------------------
# JSON
# ---------------------------------------------------------------------------

def write_json_report(result: ComparisonResult, out_path: str) -> None:
    """Grava o relatório JSON (result.to_dict, indent=2, ensure_ascii=False)."""
    _validate(result, out_path, "JSON")
    _ensure_dir(out_path)
    try:
        payload = result.to_dict()
    except Exception as exc:
        raise ValueError(
            "Falha ao serializar o resultado da comparação para JSON: %s" % exc
        )
    try:
        with open(out_path, "w", encoding="utf-8") as handle:
            json.dump(payload, handle, indent=2, ensure_ascii=False)
    except OSError as exc:
        raise ValueError(
            "Falha ao gravar o relatório JSON em '%s': %s" % (out_path, exc)
        )
    logger.info("Relatório JSON gravado em %s", out_path)


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------

_XLSX_HEADERS = [
    "ID", "Tipo", "Categoria", "Seção", "Página",
    "Texto anterior", "Texto novo", "Resumo",
]
_XLSX_WIDTHS = [8, 20, 26, 40, 10, 60, 60, 45]
_WRAP_COLUMNS = {4, 6, 7, 8}  # Seção, Texto anterior, Texto novo, Resumo (1-based)


def _xlsx_safe(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    text = _ILLEGAL_XLSX_CHARS.sub("", str(value))
    if len(text) > _XLSX_MAX_CELL:
        text = text[: _XLSX_MAX_CELL - 1] + "…"
    return text


def write_xlsx_report(result: ComparisonResult, out_path: str) -> None:
    """Grava o relatório XLSX com abas "Mudanças" e "Estatísticas"."""
    _validate(result, out_path, "XLSX")
    _ensure_dir(out_path)

    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
    )
    header_align = Alignment(vertical="center")
    wrap_align = Alignment(wrap_text=True, vertical="top")
    top_align = Alignment(vertical="top")

    workbook = Workbook()

    # --- Aba "Mudanças" ---
    ws = workbook.active
    ws.title = "Mudanças"
    for col, header in enumerate(_XLSX_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    for col, width in enumerate(_XLSX_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    changes = result.changes or []
    for row_idx, change in enumerate(changes, start=2):
        page = _page_of(change)
        values = [
            getattr(change, "id", ""),
            _type_label(getattr(change, "change_type", "")),
            _category_label(getattr(change, "category", "")),
            _breadcrumb(getattr(change, "section_path", None)),
            page if page is not None else "",
            getattr(change, "old_text", "") or "",
            getattr(change, "new_text", "") or "",
            getattr(change, "summary", "") or "",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=_xlsx_safe(value))
            cell.alignment = wrap_align if col in _WRAP_COLUMNS else top_align

    last_row = max(2, len(changes) + 1)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(_XLSX_HEADERS)), last_row)

    # --- Aba "Estatísticas" ---
    ws_stats = workbook.create_sheet("Estatísticas")
    for col, header in enumerate(("Métrica", "Valor"), start=1):
        cell = ws_stats.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    ws_stats.column_dimensions["A"].width = 42
    ws_stats.column_dimensions["B"].width = 50
    for row_idx, (metric, value) in enumerate(_stats_pairs(result), start=2):
        ws_stats.cell(row=row_idx, column=1, value=_xlsx_safe(metric)).font = Font(
            bold=False
        )
        ws_stats.cell(row=row_idx, column=2, value=_xlsx_safe(value))
    ws_stats.freeze_panes = "A2"

    try:
        workbook.save(out_path)
    except Exception as exc:
        raise ValueError(
            "Falha ao gravar o relatório XLSX em '%s': %s" % (out_path, exc)
        )
    logger.info("Relatório XLSX gravado em %s", out_path)


# ---------------------------------------------------------------------------
# HTML (arquivo único auto-contido)
# ---------------------------------------------------------------------------

def _html_payload(result: ComparisonResult) -> Dict[str, Any]:
    stats = result.stats
    changes_data: List[Dict[str, Any]] = []
    for change in result.changes or []:
        changes_data.append({
            "id": getattr(change, "id", None),
            "type": _enum_value(getattr(change, "change_type", "")),
            "category": _enum_value(getattr(change, "category", "")),
            "section": [str(s) for s in (getattr(change, "section_path", None) or [])],
            "page": _page_of(change),
            "oldText": getattr(change, "old_text", "") or "",
            "newText": getattr(change, "new_text", "") or "",
            "summary": getattr(change, "summary", "") or "",
        })
    return {
        "baseFile": os.path.basename(result.base_path) or result.base_title or "—",
        "compareFile": (
            os.path.basename(result.compare_path) or result.compare_title or "—"
        ),
        "comparedAt": result.compared_at or "",
        "durationSeconds": getattr(result, "duration_seconds", 0.0) or 0.0,
        "stats": {
            "total": stats.total_changes,
            "insertions": stats.insertions,
            "deletions": stats.deletions,
            "modifications": stats.modifications,
            "moves": stats.moves,
            "content": stats.content_changes,
            "formatting": stats.formatting_changes,
            "noise": stats.noise_changes,
            "tables": stats.table_changes,
            "images": stats.image_changes,
            "changedPages": list(stats.changed_pages or []),
            "byCategory": dict(stats.by_category or {}),
        },
        "changes": changes_data,
    }


_HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>__TITLE__</title>
<style>
:root{
  --bg:#f4f6f9; --panel:#ffffff; --text:#1c2430; --muted:#5c6675;
  --border:#dde2ea; --accent:#1a56db; --accent-soft:#e8effc;
  --ins-bg:#dbeafe; --ins-fg:#1e429f; --del-bg:#fde8e8; --del-fg:#b91c1c;
  --chip-bg:#eef1f5; --shadow:0 1px 3px rgba(16,24,40,.08);
  --bar:#1a56db;
}
html[data-theme="dark"]{
  --bg:#11151c; --panel:#1a212c; --text:#e6eaf1; --muted:#98a3b3;
  --border:#2b3546; --accent:#7aa7f7; --accent-soft:#1f2c45;
  --ins-bg:#1d3356; --ins-fg:#a9c7fb; --del-bg:#46201f; --del-fg:#f3a8a3;
  --chip-bg:#232c3a; --shadow:0 1px 3px rgba(0,0,0,.4);
  --bar:#7aa7f7;
}
*{box-sizing:border-box}
body{margin:0;font-family:-apple-system,"Segoe UI",Roboto,Helvetica,Arial,sans-serif;
  background:var(--bg);color:var(--text);font-size:14px;line-height:1.5}
header{background:var(--panel);border-bottom:1px solid var(--border);
  padding:14px 24px;display:flex;align-items:center;gap:16px;flex-wrap:wrap;
  position:sticky;top:0;z-index:10;box-shadow:var(--shadow)}
header .brand{font-size:17px;font-weight:700}
header .files{color:var(--muted);font-size:13px;min-width:0;
  overflow:hidden;text-overflow:ellipsis;white-space:nowrap;flex:1}
header .files b{color:var(--text);font-weight:600}
header .date{color:var(--muted);font-size:12px}
#themeToggle{border:1px solid var(--border);background:var(--chip-bg);color:var(--text);
  border-radius:8px;padding:6px 12px;cursor:pointer;font-size:13px}
main{max-width:1180px;margin:0 auto;padding:20px 24px 60px}
.cards{display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));
  gap:12px;margin-bottom:18px}
.card{background:var(--panel);border:1px solid var(--border);border-radius:12px;
  padding:14px 16px;box-shadow:var(--shadow)}
.card .num{font-size:24px;font-weight:700}
.card .lbl{color:var(--muted);font-size:12px;text-transform:uppercase;
  letter-spacing:.04em;margin-top:2px}
.tabs{display:flex;gap:4px;border-bottom:2px solid var(--border);
  margin-bottom:14px;flex-wrap:wrap}
.tab{border:none;background:none;color:var(--muted);padding:10px 14px;cursor:pointer;
  font-size:14px;font-weight:600;border-bottom:2px solid transparent;margin-bottom:-2px}
.tab.active{color:var(--accent);border-bottom-color:var(--accent)}
.toolbar{display:flex;gap:10px;flex-wrap:wrap;align-items:center;margin-bottom:14px}
#search{flex:1;min-width:220px;padding:8px 12px;border:1px solid var(--border);
  border-radius:8px;background:var(--panel);color:var(--text);font-size:14px}
.chips{display:flex;gap:6px;flex-wrap:wrap}
.chip{border:1px solid var(--border);background:var(--chip-bg);color:var(--muted);
  border-radius:999px;padding:4px 12px;font-size:12px;cursor:pointer;user-select:none}
.chip.on{background:var(--accent-soft);color:var(--accent);border-color:var(--accent)}
.chips-label{font-size:12px;color:var(--muted);margin-right:2px}
.count-line{color:var(--muted);font-size:13px;margin-bottom:10px}
.change{background:var(--panel);border:1px solid var(--border);border-radius:12px;
  padding:14px 16px;margin-bottom:12px;box-shadow:var(--shadow)}
.change-head{display:flex;align-items:center;gap:8px;flex-wrap:wrap;margin-bottom:8px}
.badge{border-radius:6px;padding:2px 9px;font-size:11px;font-weight:700;
  text-transform:uppercase;letter-spacing:.03em}
.badge.t-insert{background:var(--ins-bg);color:var(--ins-fg)}
.badge.t-delete{background:var(--del-bg);color:var(--del-fg)}
.badge.t-modify{background:#fef3c7;color:#92400e}
html[data-theme="dark"] .badge.t-modify{background:#453818;color:#fbd38d}
.badge.t-move,.badge.t-move_modify{background:#d1fae5;color:#065f46}
html[data-theme="dark"] .badge.t-move,html[data-theme="dark"] .badge.t-move_modify{
  background:#123c2e;color:#7ee2b8}
.badge.cat{background:var(--chip-bg);color:var(--muted);text-transform:none;font-weight:600}
.crumb{color:var(--muted);font-size:12px}
.pageref{margin-left:auto;color:var(--muted);font-size:12px;white-space:nowrap}
.summary-line{font-size:13px;color:var(--text);margin-bottom:8px}
.diff{display:grid;grid-template-columns:1fr 1fr;gap:10px}
@media(max-width:720px){.diff{grid-template-columns:1fr}}
.diff .col{border:1px solid var(--border);border-radius:8px;overflow:hidden;min-width:0}
.diff .col-title{background:var(--chip-bg);color:var(--muted);font-size:11px;
  font-weight:700;text-transform:uppercase;letter-spacing:.05em;padding:4px 10px}
.diff .col-body{padding:8px 10px;white-space:pre-wrap;word-break:break-word;font-size:13px}
.diff .empty{color:var(--muted);font-style:italic}
del{background:var(--del-bg);color:var(--del-fg);text-decoration:line-through;
  border-radius:3px;padding:0 1px}
ins{background:var(--ins-bg);color:var(--ins-fg);text-decoration:underline;
  border-radius:3px;padding:0 1px}
.nothing{background:var(--panel);border:1px dashed var(--border);border-radius:12px;
  padding:28px;text-align:center;color:var(--muted)}
#panel-stats .statbox{background:var(--panel);border:1px solid var(--border);
  border-radius:12px;padding:18px;box-shadow:var(--shadow);margin-bottom:14px}
#panel-stats h3{margin:0 0 12px;font-size:15px}
.bar-row{display:grid;grid-template-columns:220px 1fr 46px;gap:10px;
  align-items:center;margin-bottom:8px}
@media(max-width:720px){.bar-row{grid-template-columns:130px 1fr 40px}}
.bar-row .bar-label{font-size:13px;color:var(--text);overflow:hidden;
  text-overflow:ellipsis;white-space:nowrap}
.bar-track{background:var(--chip-bg);border-radius:6px;height:16px;overflow:hidden}
.bar-fill{background:var(--bar);height:100%;border-radius:6px;min-width:2px}
.bar-val{font-size:12px;color:var(--muted);text-align:right}
.kv{width:100%;border-collapse:collapse;font-size:13px}
.kv td{border-bottom:1px solid var(--border);padding:7px 8px}
.kv td:first-child{color:var(--muted);width:45%}
footer{color:var(--muted);font-size:12px;text-align:center;padding:20px}
</style>
</head>
<body>
<header>
  __LOGO__
  <div class="brand">Compare-Docs — Relatório de mudanças</div>
  <div class="files" id="hdrFiles"></div>
  <div class="date" id="hdrDate"></div>
  <button id="themeToggle" type="button">Tema escuro</button>
</header>
<main>
  <div class="cards" id="cards"></div>
  <div class="tabs" id="tabs">
    <button class="tab active" data-tab="all" type="button">Todas as mudanças</button>
    <button class="tab" data-tab="content" type="button">Só conteúdo</button>
    <button class="tab" data-tab="tabimg" type="button">Tabelas e imagens</button>
    <button class="tab" data-tab="stats" type="button">Estatísticas</button>
  </div>
  <div id="panel-list">
    <div class="toolbar">
      <input id="search" type="search" placeholder="Buscar no texto, resumo ou seção…">
    </div>
    <div class="toolbar">
      <span class="chips-label">Categoria:</span>
      <div class="chips" id="catChips"></div>
    </div>
    <div class="toolbar">
      <span class="chips-label">Tipo:</span>
      <div class="chips" id="typeChips"></div>
    </div>
    <div class="count-line" id="countLine"></div>
    <div id="list"></div>
  </div>
  <div id="panel-stats" style="display:none"></div>
</main>
<footer>Relatório gerado pelo diffAI — arquivo auto-contido, sem dependências externas.</footer>
<script type="application/json" id="cd-data">__PAYLOAD__</script>
<script>
(function(){
"use strict";
var DATA = JSON.parse(document.getElementById("cd-data").textContent);
var TYPE_LABELS = {equal:"Igual", insert:"Inserção", delete:"Exclusão",
  modify:"Modificação", move:"Movimentação", move_modify:"Mov. com modificação"};
var CAT_LABELS = {content:"Conteúdo", formatting:"Formatação",
  noise_date:"Rotineira (data)", noise_version:"Rotineira (versão)",
  noise_pagenum:"Rotineira (página)", noise_whitespace:"Rotineira (espaçamento)",
  noise_punct:"Rotineira (pontuação)", table:"Tabela", image:"Imagem",
  metadata:"Metadados"};

function esc(s){
  return String(s === null || s === undefined ? "" : s)
    .replace(/&/g,"&amp;").replace(/</g,"&lt;")
    .replace(/>/g,"&gt;").replace(/"/g,"&quot;");
}
function typeLabel(t){ return TYPE_LABELS[t] || t || "—"; }
function catLabel(c){ return CAT_LABELS[c] || c || "—"; }

/* ---------- tema ---------- */
var THEME_KEY = "comparedocs-report-theme";
var toggleBtn = document.getElementById("themeToggle");
function applyTheme(theme){
  document.documentElement.setAttribute("data-theme", theme);
  toggleBtn.textContent = theme === "dark" ? "Tema claro" : "Tema escuro";
}
var savedTheme = null;
try { savedTheme = localStorage.getItem(THEME_KEY); } catch(e) {}
var prefersDark = window.matchMedia &&
  window.matchMedia("(prefers-color-scheme: dark)").matches;
applyTheme(savedTheme || (prefersDark ? "dark" : "light"));
toggleBtn.addEventListener("click", function(){
  var next = document.documentElement.getAttribute("data-theme") === "dark"
    ? "light" : "dark";
  applyTheme(next);
  try { localStorage.setItem(THEME_KEY, next); } catch(e) {}
});

/* ---------- header ---------- */
document.getElementById("hdrFiles").innerHTML =
  "<b>" + esc(DATA.baseFile) + "</b> vs <b>" + esc(DATA.compareFile) + "</b>";
document.getElementById("hdrDate").textContent =
  DATA.comparedAt ? "Comparado em " + DATA.comparedAt : "";

/* ---------- cards ---------- */
var cardDefs = [
  ["Total de mudanças", DATA.stats.total],
  ["Conteúdo", DATA.stats.content],
  ["Rotineiras (ruído)", DATA.stats.noise],
  ["Formatação", DATA.stats.formatting],
  ["Tabelas", DATA.stats.tables],
  ["Imagens", DATA.stats.images]
];
document.getElementById("cards").innerHTML = cardDefs.map(function(cd){
  return '<div class="card"><div class="num">' + esc(cd[1]) +
    '</div><div class="lbl">' + esc(cd[0]) + "</div></div>";
}).join("");

/* ---------- diff palavra a palavra (LCS) ---------- */
function tokenize(s){
  return String(s || "").split(/(\\s+)/).filter(function(t){ return t.length > 0; });
}
function diffHtml(oldText, newText){
  var a = tokenize(oldText), b = tokenize(newText);
  if (a.length * b.length > 250000) {
    return { o: esc(oldText), n: esc(newText) };
  }
  var n = a.length, m = b.length, i, j;
  var L = [];
  for (i = 0; i <= n; i++) {
    L.push(new Array(m + 1));
    for (j = 0; j <= m; j++) { L[i][j] = 0; }
  }
  for (i = n - 1; i >= 0; i--) {
    for (j = m - 1; j >= 0; j--) {
      L[i][j] = a[i] === b[j] ? L[i+1][j+1] + 1 : Math.max(L[i+1][j], L[i][j+1]);
    }
  }
  var oh = [], nh = [];
  i = 0; j = 0;
  while (i < n && j < m) {
    if (a[i] === b[j]) { oh.push(esc(a[i])); nh.push(esc(b[j])); i++; j++; }
    else if (L[i+1][j] >= L[i][j+1]) { oh.push("<del>" + esc(a[i]) + "</del>"); i++; }
    else { nh.push("<ins>" + esc(b[j]) + "</ins>"); j++; }
  }
  while (i < n) { oh.push("<del>" + esc(a[i]) + "</del>"); i++; }
  while (j < m) { nh.push("<ins>" + esc(b[j]) + "</ins>"); j++; }
  return { o: oh.join(""), n: nh.join("") };
}

/* ---------- estado de filtros ---------- */
var activeTab = "all";
var selectedCats = {};
var selectedTypes = {};
var searchTerm = "";

function presentValues(field){
  var seen = {}, out = [];
  DATA.changes.forEach(function(c){
    var v = c[field];
    if (v && !seen[v]) { seen[v] = true; out.push(v); }
  });
  return out;
}

function buildChips(containerId, values, labelFn, selected){
  var el = document.getElementById(containerId);
  el.innerHTML = values.map(function(v){
    return '<span class="chip" data-value="' + esc(v) + '">' +
      esc(labelFn(v)) + "</span>";
  }).join("");
  el.addEventListener("click", function(ev){
    var chip = ev.target.closest(".chip");
    if (!chip) return;
    var v = chip.getAttribute("data-value");
    if (selected[v]) { delete selected[v]; chip.classList.remove("on"); }
    else { selected[v] = true; chip.classList.add("on"); }
    renderList();
  });
}
buildChips("catChips", presentValues("category"), catLabel, selectedCats);
buildChips("typeChips", presentValues("type"), typeLabel, selectedTypes);

document.getElementById("search").addEventListener("input", function(ev){
  searchTerm = ev.target.value.toLowerCase();
  renderList();
});

function tabScope(change){
  if (activeTab === "content") { return change.category === "content"; }
  if (activeTab === "tabimg") {
    return change.category === "table" || change.category === "image";
  }
  return true;
}
function hasKeys(obj){ for (var k in obj) { return true; } return false; }

function matches(change){
  if (!tabScope(change)) return false;
  if (hasKeys(selectedCats) && !selectedCats[change.category]) return false;
  if (hasKeys(selectedTypes) && !selectedTypes[change.type]) return false;
  if (searchTerm) {
    var hay = (change.oldText + " " + change.newText + " " + change.summary +
      " " + change.section.join(" ")).toLowerCase();
    if (hay.indexOf(searchTerm) === -1) return false;
  }
  return true;
}

/* ---------- render da lista ---------- */
function renderChange(c){
  var d;
  if (c.type === "insert") { d = { o: "", n: "<ins>" + esc(c.newText) + "</ins>" }; }
  else if (c.type === "delete") { d = { o: "<del>" + esc(c.oldText) + "</del>", n: "" }; }
  else { d = diffHtml(c.oldText, c.newText); }
  var crumb = c.section.length
    ? '<span class="crumb">' + esc(c.section.join(" › ")) + "</span>" : "";
  var page = (c.page !== null && c.page !== undefined)
    ? '<span class="pageref">pág. ' + esc(c.page) + "</span>"
    : '<span class="pageref"></span>';
  var summary = c.summary
    ? '<div class="summary-line">' + esc(c.summary) + "</div>" : "";
  var oldBody = d.o || '<span class="empty">— sem texto anterior —</span>';
  var newBody = d.n || '<span class="empty">— sem texto novo —</span>';
  return '<div class="change">' +
    '<div class="change-head">' +
      '<span class="badge t-' + esc(c.type) + '">' + esc(typeLabel(c.type)) + "</span>" +
      '<span class="badge cat">' + esc(catLabel(c.category)) + "</span>" +
      crumb + page +
    "</div>" + summary +
    '<div class="diff">' +
      '<div class="col"><div class="col-title">Anterior</div>' +
        '<div class="col-body">' + oldBody + "</div></div>" +
      '<div class="col"><div class="col-title">Atual</div>' +
        '<div class="col-body">' + newBody + "</div></div>" +
    "</div></div>";
}

function renderList(){
  var visible = DATA.changes.filter(matches);
  document.getElementById("countLine").textContent =
    visible.length + " de " + DATA.changes.length + " mudanças exibidas";
  var list = document.getElementById("list");
  if (!visible.length) {
    list.innerHTML = '<div class="nothing">Nenhuma mudança encontrada ' +
      "com os filtros atuais.</div>";
    return;
  }
  list.innerHTML = visible.map(renderChange).join("");
}

/* ---------- estatísticas ---------- */
function renderStats(){
  var by = DATA.stats.byCategory || {};
  var keys = Object.keys(by).sort(function(x, y){ return by[y] - by[x]; });
  var max = 1;
  keys.forEach(function(k){ if (by[k] > max) max = by[k]; });
  var bars = keys.map(function(k){
    var pct = Math.max(2, Math.round(by[k] / max * 100));
    return '<div class="bar-row"><div class="bar-label">' + esc(catLabel(k)) +
      '</div><div class="bar-track"><div class="bar-fill" style="width:' +
      pct + '%"></div></div><div class="bar-val">' + esc(by[k]) + "</div></div>";
  }).join("") || '<div class="nothing">Sem dados por categoria.</div>';
  var kvRows = [
    ["Arquivo base", DATA.baseFile],
    ["Arquivo revisado", DATA.compareFile],
    ["Data da comparação", DATA.comparedAt || "—"],
    ["Duração (s)", (Math.round(DATA.durationSeconds * 100) / 100)],
    ["Total de alterações", DATA.stats.total],
    ["Inserções", DATA.stats.insertions],
    ["Exclusões", DATA.stats.deletions],
    ["Modificações", DATA.stats.modifications],
    ["Movimentações", DATA.stats.moves],
    ["Páginas alteradas", DATA.stats.changedPages.join(", ") || "—"]
  ].map(function(r){
    return "<tr><td>" + esc(r[0]) + "</td><td>" + esc(r[1]) + "</td></tr>";
  }).join("");
  document.getElementById("panel-stats").innerHTML =
    '<div class="statbox"><h3>Mudanças por categoria</h3>' + bars + "</div>" +
    '<div class="statbox"><h3>Resumo da comparação</h3>' +
    '<table class="kv">' + kvRows + "</table></div>";
}

/* ---------- abas ---------- */
document.getElementById("tabs").addEventListener("click", function(ev){
  var tab = ev.target.closest(".tab");
  if (!tab) return;
  activeTab = tab.getAttribute("data-tab");
  var tabs = document.querySelectorAll(".tab");
  for (var i = 0; i < tabs.length; i++) { tabs[i].classList.remove("active"); }
  tab.classList.add("active");
  var isStats = activeTab === "stats";
  document.getElementById("panel-list").style.display = isStats ? "none" : "";
  document.getElementById("panel-stats").style.display = isStats ? "" : "none";
  if (isStats) { renderStats(); } else { renderList(); }
});

renderList();
})();
</script>
</body>
</html>
"""


def write_html_report(result: ComparisonResult, out_path: str) -> None:
    """Grava o relatório HTML auto-contido (CSS+JS inline, sem CDN)."""
    _validate(result, out_path, "HTML")
    _ensure_dir(out_path)

    try:
        payload = _html_payload(result)
        # "</" vira "<\/" para não fechar a tag <script> com texto do documento
        json_text = json.dumps(payload, ensure_ascii=False).replace("</", "<\\/")
    except Exception as exc:
        raise ValueError(
            "Falha ao serializar o resultado da comparação para o HTML: %s" % exc
        )

    title = "Relatório de mudanças — %s vs %s" % (
        payload["baseFile"], payload["compareFile"]
    )
    html = _HTML_TEMPLATE.replace("__TITLE__", _escape_html(title))
    html = html.replace("__PAYLOAD__", json_text)

    # Logo do escritório (plano Equipe): embutido como data URI.
    from app.branding import active_logo_data_uri

    logo_uri = active_logo_data_uri()
    logo_tag = (
        '<img src="%s" alt="" style="height:28px;max-width:160px;'
        'object-fit:contain;margin-right:12px">' % logo_uri
        if logo_uri else ""
    )
    html = html.replace("__LOGO__", logo_tag)

    try:
        with open(out_path, "w", encoding="utf-8") as handle:
            handle.write(html)
    except OSError as exc:
        raise ValueError(
            "Falha ao gravar o relatório HTML em '%s': %s" % (out_path, exc)
        )
    logger.info("Relatório HTML gravado em %s", out_path)


def _escape_html(text: str) -> str:
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )
