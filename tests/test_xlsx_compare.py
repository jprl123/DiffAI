"""Testes do pipeline XLSX: extract → compare → redline."""
from __future__ import annotations

import io
import os
import tempfile
import unittest
from typing import Tuple

from openpyxl import Workbook, load_workbook

from app.output.redline_xlsx import write_redline_xlsx
from app.xlsx.compare import compare_xlsx
from app.xlsx.extract import extract_sheets
from tests.make_samples import BASE_DIR, BUDGET_NAME, REVISED_DIR, main as make_samples


def _build_simple_pair() -> Tuple[bytes, bytes]:
    def _wb(rows: list[list[object]]) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Dados"
        for row in rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    base = _wb([
        ["Nome", "Valor"],
        ["Alice", 100],
        ["Bob", 200],
    ])
    compare = _wb([
        ["Nome", "Valor"],
        ["Alice", 100],
        ["Bob", 250],
        ["Dave", 300],
    ])
    return base, compare


class TestXlsxCompare(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        make_samples()

    def test_extract_reads_sheets(self) -> None:
        base_path = os.path.join(BASE_DIR, BUDGET_NAME)
        with open(base_path, "rb") as fh:
            sheets = extract_sheets(fh.read())
        self.assertGreaterEqual(len(sheets), 1)
        self.assertGreater(len(sheets[0].rows), 0)

    def test_compare_detects_modified_and_inserted_rows(self) -> None:
        base, compare = _build_simple_pair()
        diff = compare_xlsx(base, compare, "base.xlsx", "compare.xlsx")
        self.assertEqual(len(diff.sheets), 1)
        self.assertGreater(diff.stats.modified_cells, 0)
        self.assertGreater(diff.stats.row_add, 0)

    def test_redline_produces_valid_workbook_with_summary(self) -> None:
        base, compare = _build_simple_pair()
        out_dir = tempfile.mkdtemp(prefix="comparedocs-xlsx-")
        base_path = os.path.join(out_dir, "base.xlsx")
        compare_path = os.path.join(out_dir, "compare.xlsx")
        out_path = os.path.join(out_dir, "redline.xlsx")
        try:
            with open(base_path, "wb") as fh:
                fh.write(base)
            with open(compare_path, "wb") as fh:
                fh.write(compare)
            write_redline_xlsx(base_path, compare_path, out_path)
            self.assertTrue(os.path.isfile(out_path))
            wb = load_workbook(out_path)
            self.assertTrue(
                any(n.startswith("Compare Summary") or n == "Summary" for n in wb.sheetnames)
            )
            wb.close()
        finally:
            for name in (base_path, compare_path, out_path):
                if os.path.isfile(name):
                    os.remove(name)
            os.rmdir(out_dir)

    def test_sample_budget_redline(self) -> None:
        base_path = os.path.join(BASE_DIR, BUDGET_NAME)
        revised_path = os.path.join(REVISED_DIR, BUDGET_NAME)
        out_dir = tempfile.mkdtemp(prefix="comparedocs-xlsx-budget-")
        out_path = os.path.join(out_dir, "redline.xlsx")
        try:
            write_redline_xlsx(base_path, revised_path, out_path)
            self.assertGreater(os.path.getsize(out_path), 1000)
            wb = load_workbook(out_path)
            self.assertTrue(
                any(n.startswith("Compare Summary") or n == "Summary" for n in wb.sheetnames)
            )
            wb.close()
        finally:
            if os.path.isfile(out_path):
                os.remove(out_path)
            os.rmdir(out_dir)


if __name__ == "__main__":
    unittest.main()
